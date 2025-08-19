import os
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

# Border and style settings
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
currency_format = NamedStyle(name='currency_format')
currency_format.number_format = '#,##0.00'  # Set number format to currency without symbol, two decimal places

def apply_styles(cell, is_header=False):
    """Apply styles to the cell."""
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if is_header:
        cell.font = Font(bold=True, color="000000", name="阿里巴巴普惠体 3.0 55 Regular")  # Black font with Alibaba PuHuiTi font
        cell.fill = PatternFill("solid", fgColor="ADD8E6")  # Light blue background
    else:
        cell.font = Font(name="阿里巴巴普惠体 3.0 55 Regular")  # Set Alibaba PuHuiTi font for non-header cells
        cell.number_format = '#,##0.00'

def adjust_column_width(ws):
    """Adjust column width to fit the content."""
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = (max_length + 2) * 1.8
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

def save_excel_file(wb, directory, file_name):
    """Save the Excel workbook to the specified directory."""
    if not os.path.exists(directory):
        os.makedirs(directory)
    excel_file_path = os.path.join(directory, file_name)
    wb.save(excel_file_path)
    print(f"Excel file '{file_name}' created successfully in '{directory}'.")

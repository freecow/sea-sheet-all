#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SeaTable Excel 生成器

功能描述:
    这是一个用于从 SeaTable 生成 Excel 文件的工具，支持菜单选择不同的配置文件。
    主要功能包括：
    1. 从 SeaTable 获取数据并生成 Excel 文件
    2. 支持多个配置文件，每个配置文件可以有不同的 SeaTable API 配置
    3. 自动格式化日期和数字
    4. 智能处理百分比数值（0-1小数自动转换为0-100百分比）
    5. 使用 SUBTOTAL(109,...) 函数计算合计（只统计可见行）
    6. 支持文件合并功能
    7. 自动设置 Excel 样式和格式

配置项说明:
    每个 JSON 配置文件包含以下结构：
    {
        "date_version": "20241101",                    # 日期版本号
        "seatable_config": {                           # SeaTable 配置（可选）
            "server_url": "https://your-server.com",   # SeaTable 服务器地址
            "api_token": "your-api-token"              # SeaTable API 令牌
        },
        "excel_directories": {                         # 目录定义（可选）
            "output_dir": "/输出目录路径",             # 输出目录的引用名称
            "data_dir": "/数据目录路径"               # 数据目录的引用名称
        },
        "entries": [                                   # Excel 文件生成配置
            {
                "table_name": "表格名称",              # SeaTable 中的表格名称
                "view_name": "视图名称",               # 要使用的视图名称
                "excel_directory": "output_dir",      # 输出目录的引用名称
                "excel_file_name": "文件名.xlsx",      # 输出文件名
                "sheet_name": "工作表名称",            # 工作表名称（可选）
                "sum_columns": ["金额列1", "金额列2"], # 需要计算合计的列名列表
                "field_mapping": "all"                 # 字段映射配置（可选）
            }
        ],
        "combined_files": [                            # 文件合并配置（可选）
            {
                "output_directory": "output_dir",    # 合并文件输出目录的引用名称
                "output_file_name": "合并文件.xlsx",   # 合并后的文件名
                "include_entries": [                   # 要合并的文件列表
                    "文件1.xlsx",
                    "文件2.xlsx"
                ]
            }
        ]
    }

百分比列字段命名规则:
    程序会自动识别包含以下关键词的列名作为百分比列：
    - "比例" (如: "奖励比例S2")
    - "百分比" 
    - "percent"
    - "rate"
    - "ratio"
    - "奖励比例"
    
    对于百分比列，程序会：
    1. 自动识别 0-1 之间的小数并转换为 0-100 的百分比
    2. 保持 0-100 之间的数值不变
    3. 在 Excel 中设置为百分比格式显示

目录引用功能说明:
    1. 可以在 "excel_directories" 中定义多个目录的别名
    2. 在 "entries" 和 "combined_files" 中可以使用目录别名或直接路径
    3. 目录别名优先级：excel_directories 中定义的别名 > 直接路径
    4. 支持相对路径和绝对路径
    5. 如果使用目录别名但未在 excel_directories 中定义，程序会报错

字段映射功能说明:
    1. 支持 "all" 模式：导出所有SeaTable字段，使用原始字段名作为Excel列名
    2. 支持自定义映射：定义SeaTable字段名到Excel列名的映射关系
    3. 字段映射配置示例：
       - "all": 导出所有字段
       - {"seatable_field1": "excel_column1", "seatable_field2": "excel_column2"}: 自定义映射
    4. 如果未配置 field_mapping，默认为 "all" 模式
    5. 程序会自动验证字段映射的有效性，确保所有引用的SeaTable字段都存在

注意事项:
    1. 如果没有合并需求，请完全移除 "combined_files" 配置项，不要保留空数组
    2. 百分比列名必须包含上述关键词才能被正确识别和处理
    3. 生成的 Excel 文件会自动添加当前日期版本号（格式：@YYYYMMDD）
    4. 目录引用功能可以大大简化配置文件，避免重复的路径定义
    5. 字段映射功能可以精确控制导出的字段和Excel列名，提高数据处理的灵活性
    6. 使用自定义字段映射时，确保所有引用的SeaTable字段都存在，否则程序会报错

使用方法:
    1. 运行程序: python main-pro.py
    2. 选择配置文件
    3. 选择要生成的文件或操作

环境变量（可选）:
    如果配置文件中没有 seatable_config，可以在 .env 文件中设置：
    SEATABLE_SERVER_URL=https://your-seatable-server.com
    SEATABLE_API_TOKEN=your-api-token-here
    
    支持在 JSON 配置文件中使用环境变量占位符：
    {
        "seatable_config": {
            "server_url": "${SEATABLE_SERVER_URL}",
            "api_token": "${SEATABLE_API_TOKEN}"
        }
    }

.env 文件说明:
    1. 保留 .env 文件作为默认配置和备用方案
    2. 当 JSON 配置文件中没有 seatable_config 时，程序会自动从 .env 文件读取配置
    3. 配置优先级：JSON 配置文件中的 seatable_config > .env 文件 > 环境变量
    4. 建议将敏感信息（如 API 令牌）放在 .env 文件中，避免直接写在 JSON 配置文件里
    5. .env 文件格式示例：
       SEATABLE_SERVER_URL=https://cloud.seatable.cn
       SEATABLE_API_TOKEN=your-api-token-here
    6. 支持多环境配置：可以为不同环境创建不同的 .env 文件
    7. 支持在 JSON 配置文件中使用 ${VAR_NAME} 格式的占位符引用 .env 中的变量

Author: zhanghui
Last Modified: 2025-8-4
"""

import json
import os
import argparse
from datetime import datetime
from seatable_api import Base
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from utils.excel_utils import apply_styles, adjust_column_width, save_excel_file, currency_format
from utils.config_utils import load_and_interpolate_config
import re

# 加载 .env 文件中的环境变量
load_dotenv()

def load_config_file():
    """选择 JSON 配置文件并加载"""
    while True:
        config_files = [f for f in os.listdir('.') if f.endswith('.json')]
        if not config_files:
            print("No JSON configuration files found in the current directory.")
            return None

        print("\n请选择配置文件:")
        for i, file_name in enumerate(config_files, start=1):
            print(f"{i}. {file_name}")
        print("0. 退出")

        choice = input("请输入配置文件编号: ")
        
        if choice == '0':
            print("退出...")
            return None
        
        try:
            choice = int(choice)
            if 1 <= choice <= len(config_files):
                config_file_path = config_files[choice - 1]
                # 使用新的配置加载函数，支持环境变量插值
                config = load_and_interpolate_config(config_file_path)
                return config
            else:
                print("无效选择，请重试。")
        except ValueError:
            print("无效输入，请输入数字。")

def get_seatable_config(config):
    """从配置文件中获取 SeaTable 配置"""
    # 优先使用配置文件中的配置
    if 'seatable_config' in config:
        seatable_config = config['seatable_config']
    else:
        # 回退到环境变量
        seatable_config = {
            'server_url': os.getenv('SEATABLE_SERVER_URL'),
            'api_token': os.getenv('SEATABLE_API_TOKEN')
        }
    
    # 验证配置
    if not seatable_config.get('server_url') or not seatable_config.get('api_token'):
        raise ValueError("SeaTable 配置不完整，请检查 server_url 和 api_token")
    
    return seatable_config

def get_excel_directory(config, directory_ref):
    """根据目录引用获取实际的Excel目录路径"""
    # 如果直接是路径，直接返回
    if os.path.isabs(directory_ref) or directory_ref.startswith('./') or directory_ref.startswith('../'):
        return directory_ref
    
    # 从配置中获取目录定义
    directories = config.get('excel_directories', {})
    if directory_ref in directories:
        return directories[directory_ref]
    else:
        raise ValueError(f"未找到目录引用 '{directory_ref}'，请检查配置文件中的 excel_directories 定义")

def resolve_entries_with_directories(config):
    """解析entries配置，将目录引用替换为实际路径"""
    entries = config.get('entries', [])
    resolved_entries = []
    
    for entry in entries:
        resolved_entry = entry.copy()
        
        # 处理excel_directory
        if 'excel_directory' in resolved_entry:
            directory_ref = resolved_entry['excel_directory']
            resolved_entry['excel_directory'] = get_excel_directory(config, directory_ref)
        
        resolved_entries.append(resolved_entry)
    
    # 处理combined_files中的output_directory
    if 'combined_files' in config:
        for combined_config in config['combined_files']:
            if 'output_directory' in combined_config:
                directory_ref = combined_config['output_directory']
                combined_config['output_directory'] = get_excel_directory(config, directory_ref)
    
    return resolved_entries

# 获取当前系统日期版本
current_date_version = datetime.now().strftime('%Y%m%d')

def get_column_index_by_name(sheet, column_name):
    """根据列名获取 Excel 中的列索引，假设第1行是标题，第2行是表头"""
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=2, column=col).value == column_name:
            return col
    raise ValueError(f"Column '{column_name}' not found in Excel sheet.")

def is_date_string(value):
    """检查字符串是否为日期格式"""
    if not isinstance(value, str):
        return False
    # 匹配类似 2025-01-20T00:00:00+08:00 的格式
    date_pattern = r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}'
    return bool(re.match(date_pattern, value))

def clean_value_for_excel(value):
    """清理数据，确保Excel能正确处理"""
    if value is None:
        return ''
    
    # 处理列表类型的数据
    if isinstance(value, list):
        if len(value) == 0:
            return ''
        # 取第一个元素，如果还是列表则递归处理
        value = value[0] if isinstance(value[0], (str, int, float)) else str(value[0])
    
    if isinstance(value, str):
        # 移除控制字符
        value = ''.join(char for char in value if ord(char) >= 32 or char in '\n\r\t')
        # 移除可能导致Excel问题的字符
        value = value.replace('\x00', '').replace('\x01', '').replace('\x02', '')
        # 限制字符串长度
        if len(value) > 32000:
            value = value[:32000]
    
    return value

def is_percentage_column(column_name):
    """检查列名是否包含百分比相关关键词"""
    percentage_keywords = ['比例', '百分比', 'percent', 'rate', 'ratio', '奖励比例']
    return any(keyword in column_name for keyword in percentage_keywords)

def should_convert_to_percentage(value, column_name):
    """判断是否应该转换为百分比格式"""
    if not is_percentage_column(column_name):
        return False
    
    if value is None or value == '':
        return False
    
    try:
        # 如果是字符串，尝试转换为数字
        if isinstance(value, str):
            clean_value = value.replace('%', '').strip()
            if clean_value:
                value = float(clean_value)
            else:
                return False
        
        # 如果是数字，检查是否需要转换
        if isinstance(value, (int, float)):
            # 如果数值在 0-1 之间，且不是 0 或 1，则需要转换
            if 0 < value < 1:
                return True
            # 如果数值在 1-100 之间，已经是百分比格式
            elif 0 <= value <= 100:
                return False
            # 其他情况，不转换
            else:
                return False
    except (ValueError, TypeError):
        return False
    
    return False

def format_percentage_value(value):
    """格式化百分比数值"""
    if value is None or value == '':
        return value
    
    try:
        # 如果是字符串，尝试转换为数字
        if isinstance(value, str):
            # 移除可能的百分号
            clean_value = value.replace('%', '').strip()
            if clean_value:
                value = float(clean_value)
            else:
                return value
        
        # 如果是数字，检查数值范围来判断格式
        if isinstance(value, (int, float)):
            # 如果数值在 0-1 之间，且不是 0 或 1，则可能是小数格式
            if 0 < value < 1:
                # 转换为百分比格式（0-100）
                return value * 100
            elif 0 <= value <= 100:
                # 已经是百分比格式，保持不变
                return value
            else:
                # 其他数值，保持原样
                return value
    except (ValueError, TypeError):
        # 转换失败，保持原值
        return value
    
    return value

def format_date(value):
    """将日期字符串格式化为 yyyy-mm-dd"""
    if is_date_string(value):
        return value.split('T')[0]
    return value

def get_field_mapping(entry, all_columns):
    """获取字段映射，支持 'all' 模式和自定义映射"""
    field_mapping = entry.get('field_mapping', 'all')
    
    if field_mapping == 'all':
        # 导出所有字段，使用原始字段名作为Excel列名
        return {col: col for col in all_columns}
    elif isinstance(field_mapping, dict):
        # 使用自定义映射
        return field_mapping
    else:
        raise ValueError(f"无效的字段映射配置: {field_mapping}")

def validate_field_mapping(field_mapping, available_columns):
    """验证字段映射是否有效"""
    if field_mapping == 'all':
        return True
    
    missing_fields = []
    for seatable_field in field_mapping.keys():
        if seatable_field not in available_columns:
            missing_fields.append(seatable_field)
    
    if missing_fields:
        raise ValueError(f"以下SeaTable字段在数据中未找到: {missing_fields}")
    
    return True

def create_excel_file(entries, seatable_config):
    base = Base(seatable_config['api_token'], seatable_config['server_url'])
    base.auth()
    #base.use_api_gateway = False
    
    for entry in entries:
        table_name = entry['table_name']
        view_name = entry['view_name']
        excel_directory = entry['excel_directory']
        excel_file_name = entry['excel_file_name']
        sheet_name = entry.get('sheet_name', view_name)  # 使用 sheet_name
        sum_columns = entry['sum_columns']
        
        # 在文件名后加上系统日期版本
        file_name_without_extension, file_extension = os.path.splitext(excel_file_name)
        excel_file_name = f"{file_name_without_extension}@{current_date_version}{file_extension}"

        print(f"从 SeaTable 视图 '{view_name}' 获取数据...")
        rows = base.list_rows(table_name, view_name=view_name)
        
        if not rows:
            print(f"视图 '{view_name}' 没有找到数据，跳过...")
            continue
        
        # Filter out columns starting with _
        all_columns = [col for col in rows[0].keys() if not col.startswith('_')]
        
        # 获取字段映射
        try:
            field_mapping = get_field_mapping(entry, all_columns)
            validate_field_mapping(field_mapping, all_columns)
        except ValueError as e:
            print(f"字段映射错误: {e}")
            continue
        
        # 获取Excel列名（按映射顺序）
        excel_columns = list(field_mapping.values())
        seatable_fields = list(field_mapping.keys())
        
        # 检查哪些列不存在
        missing_columns = [col for col in sum_columns if col not in excel_columns]
        if missing_columns:
            print(f"警告: 以下列在数据中未找到: {missing_columns}")
        
        # Create Excel file
        print(f"创建 Excel 文件 '{excel_file_name}'...")
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write data with date formatting
        ws.append(excel_columns)
        for row_idx, row in enumerate(rows, start=2):
            try:
                filtered_row = []
                for seatable_field in seatable_fields:
                    value = row.get(seatable_field, '')
                    
                    # 清理数据，确保Excel能正确处理
                    value = clean_value_for_excel(value)
                    
                    # 处理百分比列
                    if is_percentage_column(seatable_field):
                        original_value = value
                        # 只有在需要转换时才进行转换
                        if should_convert_to_percentage(value, seatable_field):
                            value = format_percentage_value(value)
                    
                    filtered_row.append(format_date(value))
                ws.append(filtered_row)
            except Exception as e:
                print(f"警告: 处理第 {row_idx} 行数据时出错: {e}")
                print(f"  错误详情: 数据类型={type(row)}, 数据内容={repr(row)}")
                # 尝试写入空行或跳过
                ws.append([''] * len(excel_columns))

        # Set styles and adjust column widths
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(excel_columns)):
            for cell in row:
                is_header = cell.row == 1
                apply_styles(cell, is_header=is_header)
                
                # 检查是否是年份列，并设置为整数格式
                if isinstance(cell.value, (int, float)) and 1900 <= cell.value <= 2100:
                    cell.number_format = '0'  # 将年份设置为整数显示
                
                # 设置百分比列的格式
                if cell.row > 1 and cell.column <= len(seatable_fields):
                    seatable_field = seatable_fields[cell.column - 1]
                    if is_percentage_column(seatable_field):
                        if isinstance(cell.value, (int, float)):
                            # 对于百分比列，将数值除以100，这样Excel的百分比格式会正确显示
                            if 0 <= cell.value <= 100:
                                cell.value = cell.value / 100
                            cell.number_format = '0.00%'  # 设置为百分比格式

        adjust_column_width(ws)

        # Convert text-formatted numbers to actual number format and apply currency format
        for col in sum_columns:
            try:
                col_index = excel_columns.index(col) + 1
                col_letter = get_column_letter(col_index)
                for cell in ws[col_letter]:
                    if cell.row != 1:
                        try:
                            if cell.value is not None and str(cell.value).strip():
                                cell.value = float(cell.value)
                                cell.number_format = '#,##0.00'
                        except (ValueError, TypeError) as e:
                            print(f"警告: 单元格 {cell.coordinate} 的值 '{cell.value}' 无法转换为数字")
                            pass
            except ValueError as e:
                print(f"警告: 列 '{col}' 在数据中未找到，跳过该列的格式化")
                continue
            except Exception as e:
                print(f"错误: 处理列 '{col}' 时出错: {e}")
                continue

        # Calculate and add total row
        if sum_columns:
            total_row = ws.max_row + 1
            ws[f"A{total_row}"] = "合计"
            for col in sum_columns:
                try:
                    col_index = excel_columns.index(col) + 1
                    col_letter = get_column_letter(col_index)
                    # 使用更安全的公式写法，避免特殊字符问题
                    formula = f"=SUBTOTAL(109,{col_letter}2:{col_letter}{total_row - 1})"
                    ws[f"{col_letter}{total_row}"] = formula
                    ws[f"{col_letter}{total_row}"].style = currency_format
                except ValueError as e:
                    print(f"警告: 列 '{col}' 在数据中未找到，跳过该列的合计计算")
                    continue
                except Exception as e:
                    print(f"错误: 为列 '{col}' 添加合计公式时出错: {e}")
                    continue

            # Apply bold style to total row, similar to header
            for cell in ws[total_row]:
                apply_styles(cell, is_header=True)  # Use header style for total row

        # Remove Excel gridlines
        ws.sheet_view.showGridLines = False

        # Set header row as filter
        ws.auto_filter.ref = ws.dimensions
        
        # Save Excel file
        save_excel_file(wb, excel_directory, excel_file_name)

def combine_excel_files(combined_file_configs):
    """合并多个 Excel 文件"""
    if not combined_file_configs:
        print("没有配置合并文件，跳过合并操作。")
        return
        
    for combined_file_config in combined_file_configs:
        # 验证配置是否完整
        required_keys = ['output_directory', 'output_file_name', 'include_entries']
        missing_keys = [key for key in required_keys if key not in combined_file_config]
        
        if missing_keys:
            print(f"警告: 合并配置缺少必要字段: {missing_keys}，跳过此配置")
            continue
            
        if not combined_file_config['include_entries']:
            print("警告: 合并配置中没有包含的文件，跳过此配置")
            continue
        
        output_directory = combined_file_config['output_directory']
        output_file_name = combined_file_config['output_file_name']
        output_file_name_with_date = output_file_name.replace(".xlsx", f"@{current_date_version}.xlsx")
        
        include_entries = combined_file_config['include_entries']

        combined_wb = Workbook()
        combined_ws = combined_wb.active
        combined_ws.title = "Combined"
        combined_wb.remove(combined_ws)  # 删除默认空白表

        for entry_file in include_entries:
            file_name_with_date = entry_file.replace(".xlsx", f"@{current_date_version}.xlsx")
            entry_file_path = os.path.join(output_directory, file_name_with_date)

            if not os.path.exists(entry_file_path):
                print(f"文件 '{entry_file_path}' 未找到，跳过...")
                continue

            entry_wb = load_workbook(filename=entry_file_path)
            entry_ws = entry_wb.active
            combined_ws = combined_wb.create_sheet(title=entry_ws.title)

            # 获取源文件中的数字格式列
            number_format_columns = []
            for col in range(1, entry_ws.max_column + 1):
                cell = entry_ws.cell(row=2, column=col)  # 检查第二行（数据行）的格式
                if cell.number_format in ['#,##0.00', '#,##0', '0.00']:
                    number_format_columns.append(col)

            for row in entry_ws.iter_rows(values_only=False):
                combined_ws.append([cell.value for cell in row])

            # Apply styles to all rows
            for row in combined_ws.iter_rows(min_row=1, max_row=combined_ws.max_row):
                for cell in row:
                    # Check if this is the last row (total row)
                    is_header = cell.row == 1 or cell.row == combined_ws.max_row
                    apply_styles(cell, is_header=is_header)
                    
                    # 对数字列应用格式
                    if cell.column in number_format_columns:
                        cell.number_format = '#,##0.00'
                    # 年份特殊处理
                    elif isinstance(cell.value, (int, float)) and 1900 <= cell.value <= 2100:
                        cell.number_format = '0'

            adjust_column_width(combined_ws)
            combined_ws.sheet_view.showGridLines = False
            combined_ws.auto_filter.ref = combined_ws.dimensions

        combined_file_path = os.path.join(output_directory, output_file_name_with_date)
        combined_wb.save(combined_file_path)
        print(f"合并的 Excel 文件已保存为 {combined_file_path}")

        for entry_file in include_entries:
            file_name_with_date = entry_file.replace(".xlsx", f"@{current_date_version}.xlsx")
            entry_file_path = os.path.join(output_directory, file_name_with_date)
            if os.path.exists(entry_file_path):
                os.remove(entry_file_path)
                print(f"合并后删除源文件 '{entry_file_path}'。")

def main_menu(config):
    """第二层菜单选择"""
    while True:
        try:
            # 解析目录引用
            resolved_entries = resolve_entries_with_directories(config)
            combined_entries = config.get('combined_files', [])

            print("\n请选择要生成的 Excel 文件:")
            for i, entry in enumerate(resolved_entries, start=1):
                print(f"{i}. {entry['excel_file_name']} (视图: '{entry['view_name']}')")
            print("0. 全部生成")
            if combined_entries:
                print("c. 合并特定文件")
            print("b. 返回上级菜单")
            print("e. 退出")

            choice = input("请输入选择 (例如: 1, 2, 3, 0 表示全部, c 表示合并): ")
            
            if choice == '0':
                try:
                    seatable_config = get_seatable_config(config)
                    create_excel_file(resolved_entries, seatable_config)
                except ValueError as e:
                    print(f"配置错误: {e}")
            elif choice == 'c' and combined_entries:
                combine_excel_files(combined_entries)
            elif choice == 'b':
                return  # 返回上级菜单
            elif choice == 'e':
                print("退出...")
                exit()
            else:
                try:
                    choice = int(choice)
                    if 1 <= choice <= len(resolved_entries):
                        try:
                            seatable_config = get_seatable_config(config)
                            create_excel_file([resolved_entries[choice - 1]], seatable_config)
                        except ValueError as e:
                            print(f"配置错误: {e}")
                    else:
                        print("无效选择，请重试。")
                except ValueError:
                    print("无效输入，请输入数字。")
        except ValueError as e:
            print(f"配置解析错误: {e}")
            print("请检查配置文件中的目录引用是否正确。")
            return

def main():
    while True:
        config = load_config_file()
        if config:
            main_menu(config)
        else:
            break

if __name__ == '__main__':
    main()